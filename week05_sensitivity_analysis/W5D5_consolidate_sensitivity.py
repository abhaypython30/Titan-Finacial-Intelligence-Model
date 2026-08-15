"""Titan Company - Week5 Day5: Consolidate Sensitivity Findings"""
import os, logging
import pandas as pd
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.join(SCRIPT_DIR, "..")
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
EXCEL_PATH = os.path.join(DATA_DIR, "Titan project.xlsx")
OUTPUT_CSV = os.path.join(DATA_DIR, "titan_week5_sensitivity_master_summary.csv")


def load_all_days():
    files = {
        "wacc_tg_grid": "titan_sensitivity_wacc_tg.csv",
        "growth_margin_grid": "titan_sensitivity_growth_margin.csv",
        "montecarlo_summary": "titan_montecarlo_summary.csv",
        "driver_ranking": "titan_driver_ranking.csv",
    }
    loaded = {}
    for key, filename in files.items():
        path = os.path.join(DATA_DIR, filename)
        if os.path.exists(path):
            loaded[key] = pd.read_csv(path)
        else:
            log.warning(f"{filename} not found.")
            loaded[key] = None
    return loaded


def build_master_summary(loaded):
    lines = []
    if loaded["wacc_tg_grid"] is not None:
        vals = [v for v in loaded["wacc_tg_grid"].drop(columns=["wacc"]).values.flatten() if pd.notna(v)]
        lines.append({"finding": "WACC x Terminal Growth spread",
                       "detail": f"Rs.{min(vals):.2f} to Rs.{max(vals):.2f} ({max(vals)/min(vals):.1f}x)"})

    if loaded["growth_margin_grid"] is not None:
        cols = [c for c in loaded["growth_margin_grid"].columns if c.startswith("margin_")]
        vals = [v for v in loaded["growth_margin_grid"][cols].values.flatten() if pd.notna(v)]
        neg = sum(1 for v in vals if v < 0)
        lines.append({"finding": "Growth x Margin range",
                       "detail": f"Rs.{min(vals):.2f} to Rs.{max(vals):.2f} ({neg} negative cells)"})

    if loaded["montecarlo_summary"] is not None:
        mc = loaded["montecarlo_summary"].iloc[0]
        lines.append({"finding": "Monte Carlo median value/share",
                       "detail": f"Rs.{mc['median']:.2f} (P5={mc['p5']:.2f}, P95={mc['p95']:.2f}), "
                                 f"{mc['probability_negative_equity_pct']:.1f}% negative-equity probability"})

    if loaded["driver_ranking"] is not None:
        top = loaded["driver_ranking"].iloc[0]
        lines.append({"finding": "Most impactful assumption",
                       "detail": f"{top['driver']} - Rs.{top['absolute_swing_rs']:.2f} "
                                 f"({top['swing_as_pct_of_base_value']:.1f}% of base)"})

    summary_df = pd.DataFrame(lines)
    log.info("Week5 master findings:")
    for _, row in summary_df.iterrows():
        log.info(f"  {row['finding']}: {row['detail']}")
    return summary_df


def write_to_excel(summary_df):
    if not os.path.exists(EXCEL_PATH):
        log.warning("Excel file not found - skipping write.")
        return
    wb = load_workbook(EXCEL_PATH)
    if "Sensitivity_Summary" in wb.sheetnames:
        del wb["Sensitivity_Summary"]
    ws = wb.create_sheet("Sensitivity_Summary")
    ws["A1"] = "Week 5 - Sensitivity Analysis Summary (Titan)"
    ws["A3"], ws["B3"] = "Finding", "Detail"
    for i, row in summary_df.iterrows():
        ws[f"A{4+i}"] = row["finding"]
        ws[f"B{4+i}"] = row["detail"]
    wb.save(EXCEL_PATH)
    log.info(f"Wrote Sensitivity_Summary sheet to {EXCEL_PATH}")


def main():
    loaded = load_all_days()
    summary_df = build_master_summary(loaded)
    summary_df.to_csv(OUTPUT_CSV, index=False)
    log.info(f"Saved to {OUTPUT_CSV}")
    write_to_excel(summary_df)


if __name__ == "__main__":
    main()