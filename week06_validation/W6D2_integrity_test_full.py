"""Titan Company - Week 6 Day 2: Full Integrity Test (adapted from Tata's 21-check suite)"""

import os
import re
import logging
import pandas as pd
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.join(SCRIPT_DIR, "..")
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
WEEK4_DIR = os.path.join(PROJECT_ROOT, "week04_financial_model")
EXCEL_PATH = os.path.join(DATA_DIR, "Titan project.xlsx")

results = []


def record(day, name, passed, detail):
    results.append((day, name, passed, detail))
    (log.info if passed else log.warning)(f"[{'PASS' if passed else 'FAIL'}] {day} - {name}: {detail}")


def read_excel():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Assumptions"]
    return {
        "wacc": ws["C9"].value,
        "rev_bear": ws["C13"].value, "rev_base": ws["C12"].value, "rev_bull": ws["C14"].value,
        "mgn_bear": ws["C18"].value, "mgn_base": ws["C17"].value, "mgn_bull": ws["C19"].value,
        "tg": ws["C22"].value, "dep_pct": ws["C23"].value,
        "capex_pct": ws["C24"].value, "tax_rate": ws["C25"].value,
    }


def close(a, b, tol=0.05):
    return a is not None and b is not None and abs(a - b) <= tol


def check_w4d1(excel):
    path = os.path.join(DATA_DIR, "titan_assumptions_derived.csv")
    if not os.path.exists(path):
        record("W4D1", "WACC present", False, "titan_assumptions_derived.csv missing")
        return
    df = pd.read_csv(path)
    record("W4D1", "WACC matches CSV", close(excel["wacc"], df["wacc"].iloc[0]),
           f"Excel={excel['wacc']}%, CSV={df['wacc'].iloc[0]}%")


def check_w4d2(excel):
    path = os.path.join(DATA_DIR, "titan_revenue_drivers.csv")
    if not os.path.exists(path):
        record("W4D2", "Revenue scenarios match", False, "file missing")
        return
    row = pd.read_csv(path).iloc[0]
    ok = close(excel["rev_bear"], row["final_bear"]) and close(excel["rev_base"], row["final_base"]) and close(excel["rev_bull"], row["final_bull"])
    record("W4D2", "Excel revenue matches locked CSV", ok, f"{excel['rev_bear']}/{excel['rev_base']}/{excel['rev_bull']}")


def check_w4d3(excel):
    path = os.path.join(DATA_DIR, "titan_margin_drivers.csv")
    if not os.path.exists(path):
        record("W4D3", "Margin scenarios match", False, "file missing")
        return
    row = pd.read_csv(path).iloc[0]
    ok = close(excel["mgn_bear"], row["final_bear"]) and close(excel["mgn_base"], row["final_base"]) and close(excel["mgn_bull"], row["final_bull"])
    record("W4D3", "Excel margin matches locked CSV", ok, f"{excel['mgn_bear']}/{excel['mgn_base']}/{excel['mgn_bull']}")


def check_w4d4(excel):
    path = os.path.join(DATA_DIR, "titan_capex_dep_terminal.csv")
    if not os.path.exists(path):
        record("W4D4", "Other Inputs match", False, "file missing")
        return
    row = pd.read_csv(path).iloc[0]
    ok = (close(excel["capex_pct"], row["capex_pct_sales"]) and close(excel["dep_pct"], row["depreciation_pct_sales"])
          and close(excel["tax_rate"], row["effective_tax_rate"]) and close(excel["tg"], row["terminal_growth_rate"]))
    record("W4D4", "Excel Other Inputs match locked CSV", ok, "checked")
    record("W4D4", "Terminal growth < WACC", excel["tg"] < excel["wacc"], f"{excel['tg']}% < {excel['wacc']}%")
    record("W4D4", "CAPEX% > Depreciation%", excel["capex_pct"] > excel["dep_pct"], f"{excel['capex_pct']}% > {excel['dep_pct']}%")


def check_w4d5():
    path = os.path.join(DATA_DIR, "titan_dcf_valuation.csv")
    if not os.path.exists(path):
        record("W4D5", "DCF valid", False, "file missing")
        return
    df = pd.read_csv(path)
    record("W4D5", "3 scenarios present", set(df["scenario"]) == {"bear", "base", "bull"}, str(set(df["scenario"])))
    bear_v = df[df["scenario"] == "bear"]["value_per_share"].iloc[0]
    base_v = df[df["scenario"] == "base"]["value_per_share"].iloc[0]
    bull_v = df[df["scenario"] == "bull"]["value_per_share"].iloc[0]
    record("W4D5", "Bear<Base<Bull", bear_v < base_v < bull_v, f"{bear_v:.2f}<{base_v:.2f}<{bull_v:.2f}")
    record("W4D5", "EV positive all scenarios", (df["enterprise_value_cr"] > 0).all(), "checked")


def check_w5d1():
    path = os.path.join(DATA_DIR, "titan_sensitivity_wacc_tg.csv")
    if not os.path.exists(path):
        record("W5D1", "Grid monotonic", False, "file missing")
        return
    grid = pd.read_csv(path).set_index("wacc")
    wacc_mono = all(grid[c].dropna().is_monotonic_decreasing for c in grid.columns)
    record("W5D1", "Value decreases as WACC increases", wacc_mono, "checked")


def check_w5d3():
    rp = os.path.join(DATA_DIR, "titan_montecarlo_results.csv")
    sp = os.path.join(DATA_DIR, "titan_montecarlo_summary.csv")
    if not (os.path.exists(rp) and os.path.exists(sp)):
        record("W5D3", "MC probability matches raw", False, "file missing")
        return
    raw = pd.read_csv(rp)
    summary = pd.read_csv(sp)
    recomputed = (raw["value_per_share"] < 0).mean() * 100
    reported = summary["probability_negative_equity_pct"].iloc[0]
    record("W5D3", "Reported probability matches raw recompute", close(recomputed, reported, tol=0.1),
           f"Reported={reported:.2f}%, Recomputed={recomputed:.2f}%")


def check_w5d4():
    path = os.path.join(DATA_DIR, "titan_driver_ranking.csv")
    if not os.path.exists(path):
        record("W5D4", "Ranking sorted", False, "file missing")
        return
    df = pd.read_csv(path).sort_values("rank")
    record("W5D4", "Sorted descending by swing", df["absolute_swing_rs"].is_monotonic_decreasing,
           str(df["driver"].tolist()))


def check_no_regression():
    day1_path = os.path.join(WEEK4_DIR, "W4D1_assumptions.py")
    if not os.path.exists(day1_path):
        record("Regression", "Day1 doesn't overwrite scenario cells", False, "file not found")
        return
    content = open(day1_path, "r", encoding="utf-8", errors="ignore").read()
    dangerous = re.search(r'ws\[.*[Cc](1[2-9])', content)
    record("Regression", "Day1 doesn't overwrite C12-C19", dangerous is None,
           "clean" if dangerous is None else f"found: {dangerous.group()}")


def main():
    excel = read_excel()
    check_w4d1(excel)
    check_w4d2(excel)
    check_w4d3(excel)
    check_w4d4(excel)
    check_w4d5()
    check_w5d1()
    check_w5d3()
    check_w5d4()
    check_no_regression()

    passed = sum(1 for r in results if r[2])
    total = len(results)
    log.info(f"\n=== INTEGRITY TEST: {passed}/{total} passed ===")
    for day, name, p, detail in results:
        log.info(f"  [{'PASS' if p else 'FAIL'}] {day} - {name}")
    if passed < total:
        log.warning(f"{total-passed} check(s) failed - review before Day3.")


if __name__ == "__main__":
    main()