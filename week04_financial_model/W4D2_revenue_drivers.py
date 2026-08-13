"""
Titan Company - Week 4, Day 2: Revenue Growth Drivers

"""

import logging
import pandas as pd

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

PIPELINE_CSV = "data/titan_pipeline_new.csv"
EXCEL_PATH = "data/Titan project.xlsx"
ASSUMPTIONS_SHEET = "Assumptions"
OUTPUT_CSV = "data/titan_revenue_drivers.csv"

COVID_YEARS = [2020, 2021]


def load_pipeline():
    df = pd.read_csv(PIPELINE_CSV)
    df = df.sort_values("year").reset_index(drop=True)
    return df


def compute_growth(df):
    df["growth_pct"] = df["sales"].pct_change() * 100
    return df


def full_stats_table(growth):
    stats = {
        "mean": growth.mean(), "std": growth.std(),
        "min": growth.min(), "max": growth.max(),
        "p10": growth.quantile(0.10), "p25": growth.quantile(0.25),
        "p50_median": growth.quantile(0.50), "p75": growth.quantile(0.75),
        "p90": growth.quantile(0.90),
    }
    log.info("Full growth distribution:")
    for k, v in stats.items():
        log.info(f"  {k}: {v:.2f}%")
    return stats


def covid_decision(df):
    growth = df.dropna(subset=["growth_pct"])
    non_covid = growth[~growth["year"].isin(COVID_YEARS)]

    # FIXED: 'last 3 years' comparison doesn't work for Titan's FY17-26
    # window - 2024-2026 never contained a COVID year, so that comparison
    # would show almost no difference regardless of the decision. Instead,
    # compare the FULL-DISTRIBUTION percentiles directly - this is the
    # metric that actually determines the locked Bear/Base/Bull values,
    # so it's the one that should drive the decision.
    incl_base, incl_bear, incl_bull = (
        growth["growth_pct"].quantile(0.50),
        growth["growth_pct"].quantile(0.25),
        growth["growth_pct"].quantile(0.75),
    )
    excl_base, excl_bear, excl_bull = (
        non_covid["growth_pct"].quantile(0.50),
        non_covid["growth_pct"].quantile(0.25),
        non_covid["growth_pct"].quantile(0.75),
    )

    log.info(f"Full-distribution comparison (this is the metric that matters):")
    log.info(f"  Base (P50): incl.={incl_base:.2f}% vs excl.={excl_base:.2f}% "
              f"(diff {abs(incl_base-excl_base):.2f}pp)")
    log.info(f"  Bear (P25): incl.={incl_bear:.2f}% vs excl.={excl_bear:.2f}% "
              f"(diff {abs(incl_bear-excl_bear):.2f}pp)")
    log.info(f"  Bull (P75): incl.={incl_bull:.2f}% vs excl.={excl_bull:.2f}% "
              f"(diff {abs(incl_bull-excl_bull):.2f}pp)")

    # DECISION: Given the COVID trough (4.62% per Week2) is a genuine
    # demand-side anomaly, and Recovery already overshoots Pre-COVID
    # levels (32.61% vs 22.13%), excluding COVID years avoids anchoring
    # the Base case to an artificially depressed year.
    # for a different, Titan-specific reason
    # (demand shock during lockdowns, not a margin effect).
    decision = "excluded"
    decided_growth = non_covid["growth_pct"]

    rationale = (
        f"COVID years (2020-2021) EXCLUDED. Full-distribution Base (P50) shifts from "
        f"{incl_base:.2f}% (incl.) to {excl_base:.2f}% (excl.) - this is the real, "
        f"decision-relevant comparison, not a 'last 3 years' average which doesn't "
        f"overlap COVID for this year range. Rationale: Titan's COVID impact was "
        f"specifically on GROWTH (demand-side lockdown effect on retail footfall), "
        f"not margin - Recovery already overshoots Pre-COVID growth levels, "
        f"supporting exclusion of the COVID trough as non-representative."
    )
    log.warning(rationale)

    return decided_growth, decision, rationale


def lock_final_scenarios(decided_growth):
    final = {
        "base": decided_growth.quantile(0.50),
        "bear": decided_growth.quantile(0.25),
        "bull": decided_growth.quantile(0.75),
    }
    log.info(f"Final locked growth - Base: {final['base']:.2f}% | "
              f"Bear: {final['bear']:.2f}% | Bull: {final['bull']:.2f}%")
    return final


def write_to_excel(final):
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH)
    ws = wb[ASSUMPTIONS_SHEET]
    for col in ["C", "D", "E", "F", "G"]:
        ws[f"{col}12"] = round(final["base"], 2)
        ws[f"{col}13"] = round(final["bear"], 2)
        ws[f"{col}14"] = round(final["bull"], 2)
    wb.save(EXCEL_PATH)
    log.info("Overwrote C12:G14 with final Base/Bear/Bull growth")


def main():
    df = load_pipeline()
    df = compute_growth(df)

    stats = full_stats_table(df["growth_pct"].dropna())
    decided_growth, decision, rationale = covid_decision(df)
    final = lock_final_scenarios(decided_growth)

    write_to_excel(final)

    out = pd.DataFrame([{
        **stats, "covid_decision": decision, "rationale": rationale,
        "final_base": final["base"], "final_bear": final["bear"], "final_bull": final["bull"],
    }])
    out.to_csv(OUTPUT_CSV, index=False)
    log.info(f"Saved to {OUTPUT_CSV}")


if __name__ == "__main__":
    main()