"""
Titan Company - Week 4, Day 3: EBITDA Margin & Cost Structure Drivers


"""

import logging
import pandas as pd

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

PIPELINE_CSV = "data/titan_pipeline_new.csv"
EXCEL_PATH = "data/Titan project.xlsx"
ASSUMPTIONS_SHEET = "Assumptions"
OUTPUT_CSV = "data/titan_margin_drivers.csv"

COVID_YEARS = [2020, 2021]
COST_LINES = ["raw_material_cost", "power_fuel", "other_mfr_exp", "employee_cost",
              "selling_admin", "other_expenses"]


def load_pipeline():
    df = pd.read_csv(PIPELINE_CSV)
    df = df.sort_values("year").reset_index(drop=True)
    return df


def compute_margin_and_cost_ratios(df):
    # Uses EXISTING verified 'ebitda' column - confirmed against Titan's
    # own HistoricalFS tab in Week1 (exact match at FY24: Rs.5,292cr).
    df["ebitda_margin_pct"] = (df["ebitda"] / df["sales"]) * 100

    for line in COST_LINES:
        df[f"{line}_pct_sales"] = (df[line] / df["sales"]) * 100

    log.info(f"Margin range: {df['ebitda_margin_pct'].min():.2f}% to "
              f"{df['ebitda_margin_pct'].max():.2f}% - cross-check against "
              f"Week3's SQL-derived benchmark.")
    return df


def find_main_margin_driver(df):
    margin_change = df["ebitda_margin_pct"].diff()
    correlations = {}
    for line in COST_LINES:
        cost_change = df[f"{line}_pct_sales"].diff()
        correlations[line] = margin_change.corr(cost_change)

    log.info("Correlation of each cost ratio's YoY change vs EBITDA margin's YoY change:")
    for line, corr in sorted(correlations.items(), key=lambda x: x[1]):
        log.info(f"  {line}: {corr:.2f}")

    main_driver = min(correlations, key=correlations.get)
    log.info(f"Main margin driver identified: '{main_driver}' "
              f"(correlation {correlations[main_driver]:.2f})")

    if main_driver != "raw_material_cost":
        log.warning(f"NOTE: prediction was 'raw_material_cost' (gold cost) "
                    f"would dominate - actual result is '{main_driver}'. "
                    f"Report the ACTUAL finding, don't force the prediction.")

    return correlations, main_driver


def margin_covid_decision(df):
    margin = df.dropna(subset=["ebitda_margin_pct"])
    non_covid = margin[~margin["year"].isin(COVID_YEARS)]

    # FIXED: same issue as Day2 - 'last 3 years' doesn't overlap COVID for
    # Titan's year range. Compare full-distribution percentiles instead.
    incl_base = margin["ebitda_margin_pct"].quantile(0.50)
    excl_base = non_covid["ebitda_margin_pct"].quantile(0.50)
    diff = abs(incl_base - excl_base)

    log.info(f"Full-distribution margin comparison: Base (P50) incl.={incl_base:.2f}% "
              f"vs excl.={excl_base:.2f}% (diff {diff:.2f}pp)")

    # Per Week2: margin barely moved during COVID (9.69% -> 9.83%) -
    # expect this diff to be SMALL, unlike the growth diff in Day2.
    if diff < 1.0:
        log.info("Diff is small - consistent with Week2's finding that "
                  "COVID did not meaningfully affect margin. Excluding or "
                  "including COVID years makes little practical difference "
                  "here, unlike Day2's revenue decision.")

    decision = "excluded"  # kept consistent with Day2 for methodology uniformity
    decided_margin = non_covid["ebitda_margin_pct"]
    rationale = (f"COVID years excluded for consistency with Day2, though impact "
                f"on margin specifically was minimal (Base P50 diff={diff:.2f}pp) - "
                f"unlike the meaningful growth-side impact.")
    log.warning(rationale)

    return decided_margin, decision, rationale


def lock_final_scenarios(decided_margin):
    final = {
        "base": decided_margin.quantile(0.50),
        "bear": decided_margin.quantile(0.25),
        "bull": decided_margin.quantile(0.75),
    }
    log.info(f"Final locked margin - Base: {final['base']:.2f}% | "
              f"Bear: {final['bear']:.2f}% | Bull: {final['bull']:.2f}%")
    return final


def write_to_excel(final):
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH)
    ws = wb[ASSUMPTIONS_SHEET]
    for col in ["C", "D", "E", "F", "G"]:
        ws[f"{col}17"] = round(final["base"], 2)
        ws[f"{col}18"] = round(final["bear"], 2)
        ws[f"{col}19"] = round(final["bull"], 2)
    wb.save(EXCEL_PATH)
    log.info("Overwrote C17:G19 with final Base/Bear/Bull margin")


def main():
    df = load_pipeline()
    df = compute_margin_and_cost_ratios(df)

    correlations, main_driver = find_main_margin_driver(df)
    decided_margin, decision, rationale = margin_covid_decision(df)
    final = lock_final_scenarios(decided_margin)

    write_to_excel(final)

    out = pd.DataFrame([{
        "main_margin_driver": main_driver,
        **{f"corr_{k}": v for k, v in correlations.items()},
        "covid_decision": decision, "rationale": rationale,
        "final_base": final["base"], "final_bear": final["bear"], "final_bull": final["bull"],
    }])
    out.to_csv(OUTPUT_CSV, index=False)
    log.info(f"Saved to {OUTPUT_CSV}")


if __name__ == "__main__":
    main()