"""
Titan Company - Week 4, Day 4: CAPEX%, Depreciation%, Tax Rate, Terminal Growth

"""

import logging
import pandas as pd

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

PIPELINE_CSV = "data/titan_pipeline_new.csv"
ASSUMPTIONS_DERIVED_CSV = "data/titan_assumptions_derived.csv"
EXCEL_PATH = "data/Titan project.xlsx"
ASSUMPTIONS_SHEET = "Assumptions"
OUTPUT_CSV = "data/titan_capex_dep_terminal.csv"

COVID_YEARS = [2020, 2021]
CELL_TERMINAL_GROWTH = "C22"
CELL_DEPRECIATION_PCT = "C23"
CELL_CAPEX_PCT = "C24"
CELL_EFFECTIVE_TAX = "C25"

TERMINAL_GROWTH_MAX_BUFFER_BELOW_WACC = 2.0
CONVENTIONAL_TERMINAL_GROWTH = 4.0  # same India nominal GDP proxy as Tata


def load_pipeline():
    df = pd.read_csv(PIPELINE_CSV)
    df = df.sort_values("year").reset_index(drop=True)
    return df


def capex_pct_locked(df):
    df["capex_pct_sales"] = df["capex"] / df["sales"] * 100
    non_covid = df[~df["year"].isin(COVID_YEARS)]
    last3 = non_covid["capex_pct_sales"].tail(3).mean()
    log.info(f"CAPEX % of sales (last-3yr, COVID-excl.): {last3:.2f}%")
    return last3


def depreciation_pct_locked(df):
    df["dep_pct_sales"] = df["depreciation"] / df["sales"] * 100
    non_covid = df[~df["year"].isin(COVID_YEARS)]
    last3 = non_covid["dep_pct_sales"].tail(3).mean()
    log.info(f"Depreciation % of sales (last-3yr, COVID-excl.): {last3:.2f}%")
    return last3


def effective_tax_reused(assumptions_csv):
    day1 = pd.read_csv(assumptions_csv)
    tax_rate = day1["tax_rate_implied"].iloc[0]
    log.info(f"Effective tax rate (reused from Day1): {tax_rate:.2f}%")
    return tax_rate


def terminal_growth_with_flow_check(wacc):
    max_allowed = wacc - TERMINAL_GROWTH_MAX_BUFFER_BELOW_WACC
    terminal_growth = min(CONVENTIONAL_TERMINAL_GROWTH, max_allowed)

    if CONVENTIONAL_TERMINAL_GROWTH > max_allowed:
        log.warning(f"Conventional terminal growth exceeds safe margin below "
                    f"WACC ({wacc:.2f}%) - capped to {terminal_growth:.2f}%.")
    else:
        log.info(f"Terminal growth {terminal_growth:.2f}% safely below "
                  f"WACC {wacc:.2f}% - flow check passed.")

    assert terminal_growth < wacc, (
        f"CRITICAL: terminal growth ({terminal_growth:.2f}%) >= WACC ({wacc:.2f}%).")
    return terminal_growth


def capex_vs_depreciation_check(capex_pct, dep_pct):
    if capex_pct <= dep_pct:
        log.warning(f"CAPEX% ({capex_pct:.2f}%) <= Depreciation% ({dep_pct:.2f}%) - "
                    f"shrinking net asset base - check if expected for Titan "
                    f"(possible if store footprint growth has slowed) or a data issue.")
    else:
        log.info(f"CAPEX% ({capex_pct:.2f}%) > Depreciation% ({dep_pct:.2f}%) - "
                  f"expanding asset base - passed.")


def write_to_excel(terminal_growth, dep_pct, capex_pct, tax_rate):
    from openpyxl import load_workbook
    wb = load_workbook(EXCEL_PATH)
    ws = wb[ASSUMPTIONS_SHEET]
    ws[CELL_TERMINAL_GROWTH] = round(terminal_growth, 2)
    ws[CELL_DEPRECIATION_PCT] = round(dep_pct, 2)
    ws[CELL_CAPEX_PCT] = round(capex_pct, 2)
    ws[CELL_EFFECTIVE_TAX] = round(tax_rate, 2)
    wb.save(EXCEL_PATH)
    log.info("Wrote terminal growth, depreciation%, capex%, effective tax to 'Other Inputs'.")


def main():
    df = load_pipeline()

    day1 = pd.read_csv(ASSUMPTIONS_DERIVED_CSV)
    wacc = day1["wacc"].iloc[0]
    log.info(f"WACC from Day1 (reused): {wacc:.2f}%")

    capex_pct = capex_pct_locked(df)
    dep_pct = depreciation_pct_locked(df)
    tax_rate = effective_tax_reused(ASSUMPTIONS_DERIVED_CSV)
    terminal_growth = terminal_growth_with_flow_check(wacc)

    capex_vs_depreciation_check(capex_pct, dep_pct)

    write_to_excel(terminal_growth, dep_pct, capex_pct, tax_rate)

    out = pd.DataFrame([{
        "capex_pct_sales": capex_pct, "depreciation_pct_sales": dep_pct,
        "effective_tax_rate": tax_rate, "terminal_growth_rate": terminal_growth,
        "wacc_reference": wacc,
    }])
    out.to_csv(OUTPUT_CSV, index=False)
    log.info(f"Saved to {OUTPUT_CSV}")


if __name__ == "__main__":
    main()