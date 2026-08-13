"""
Titan Company - Week 4, Day 5: FY27-31 Forecast & DCF Valuation

"""

import logging
import pandas as pd
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

PIPELINE_CSV = "data/titan_pipeline_new.csv"
EXCEL_PATH = "data/Titan project.xlsx"
ASSUMPTIONS_SHEET = "Assumptions"
OUTPUT_CSV = "data/titan_dcf_valuation.csv"

FORECAST_YEARS = [2027, 2028, 2029, 2030, 2031]
FACE_VALUE_RS = 1  # confirmed via Week1 verification, NOT Tata's Rs.2


def get_column_with_fallback(df, preferred_col, fallback_col, ratio_name):
    if preferred_col in df.columns:
        return df[preferred_col]
    elif fallback_col in df.columns:
        log.info(f"[{ratio_name}] '{preferred_col}' not found - using fallback '{fallback_col}'")
        return df[fallback_col]
    else:
        raise KeyError(f"[{ratio_name}] Neither '{preferred_col}' nor '{fallback_col}' found.")


def preflight_cross_check():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[ASSUMPTIONS_SHEET]

    checks = {
        "wacc_at_C9": ws["C9"].value,
        "revenue_base_C12": ws["C12"].value, "revenue_bear_C13": ws["C13"].value,
        "revenue_bull_C14": ws["C14"].value,
        "margin_base_C17": ws["C17"].value, "margin_bear_C18": ws["C18"].value,
        "margin_bull_C19": ws["C19"].value,
        "terminal_growth_C22": ws["C22"].value, "depreciation_pct_C23": ws["C23"].value,
        "capex_pct_C24": ws["C24"].value, "effective_tax_C25": ws["C25"].value,
    }

    log.info("Preflight - live Excel read:")
    for k, v in checks.items():
        log.info(f"  {k}: {v}")

    failures = []
    if checks["wacc_at_C9"] is None:
        failures.append("WACC missing at C9.")
    if not (checks["revenue_bear_C13"] < checks["revenue_base_C12"] < checks["revenue_bull_C14"]):
        failures.append("Revenue Bear<Base<Bull ordering violated.")
    if not (checks["margin_bear_C18"] < checks["margin_base_C17"] < checks["margin_bull_C19"]):
        failures.append("Margin Bear<Base<Bull ordering violated.")
    if checks["terminal_growth_C22"] >= checks["wacc_at_C9"]:
        failures.append("Terminal growth >= WACC.")
    if checks["capex_pct_C24"] <= checks["depreciation_pct_C23"]:
        failures.append("CAPEX% <= Depreciation% - verify if intended for Titan.")

    if failures:
        log.warning("PREFLIGHT FAILED:")
        for f in failures:
            log.warning(f"  - {f}")
        raise SystemExit("Preflight cross-check failed.")

    log.info("Preflight cross-check PASSED.")
    return checks


def project_financials(df, checks, scenario, terminal_growth):
    growth = {"bear": checks["revenue_bear_C13"], "base": checks["revenue_base_C12"],
              "bull": checks["revenue_bull_C14"]}[scenario] / 100
    margin = {"bear": checks["margin_bear_C18"], "base": checks["margin_base_C17"],
              "bull": checks["margin_bull_C19"]}[scenario] / 100
    tg_dec = terminal_growth / 100

    dep_pct = checks["depreciation_pct_C23"] / 100
    capex_pct = checks["capex_pct_C24"] / 100
    tax_rate = checks["effective_tax_C25"] / 100

    last_sales = df["sales"].iloc[-1]
    payables = get_column_with_fallback(df, "payables", "other_liabilities", "nwc")
    last_nwc = (df["debtors"] + df["inventory"] + df["cash_operating"] - payables).iloc[-1]
    nwc_pct_sales = last_nwc / last_sales

    log.info(f"[{scenario}] NWC is {nwc_pct_sales:.1%} of sales - given Titan's "
              f"high working-capital intensity (verified in Week2), holding this "
              f"ratio flat while growth compounds at {growth:.1%} for 5 years "
              f"without tapering would produce an unrealistic cash drag. Growth "
              f"is therefore tapered toward terminal growth below.")

    rows = []
    prev_sales, prev_nwc = last_sales, last_nwc
    n_years = len(FORECAST_YEARS)

    for i, year in enumerate(FORECAST_YEARS):
        # TAPER: linearly fade from the locked scenario growth rate in
        # year 1 toward terminal growth by the final forecast year - a
        # standard DCF refinement, not a workaround. Prevents 5 straight
        # years of un-moderated high growth from producing an implausible
        # terminal-value blowup when combined with high NWC intensity.
        step = i / (n_years - 1) if n_years > 1 else 0
        current_growth = growth - (growth - tg_dec) * step

        sales = prev_sales * (1 + current_growth)
        ebitda = sales * margin
        depreciation = sales * dep_pct
        ebit = ebitda - depreciation
        tax = max(ebit, 0) * tax_rate
        nopat = ebit - tax
        capex = sales * capex_pct
        nwc = sales * nwc_pct_sales
        d_nwc = nwc - prev_nwc
        fcff = nopat + depreciation - capex - d_nwc

        rows.append({"year": year, "scenario": scenario, "applied_growth": current_growth * 100,
                     "sales": sales, "ebitda": ebitda, "depreciation": depreciation,
                     "ebit": ebit, "tax": tax, "nopat": nopat,
                     "capex": capex, "nwc": nwc, "d_nwc": d_nwc, "fcff": fcff})
        prev_sales, prev_nwc = sales, nwc

    return pd.DataFrame(rows)


def discount_and_value(forecast_df, wacc, terminal_growth, df, checks):
    wacc_dec, tg_dec = wacc / 100, terminal_growth / 100

    forecast_df["discount_factor"] = [1 / (1 + wacc_dec) ** (i + 1) for i in range(len(forecast_df))]
    forecast_df["pv_fcff"] = forecast_df["fcff"] * forecast_df["discount_factor"]

    final_year_fcff = forecast_df["fcff"].iloc[-1]
    terminal_value = final_year_fcff * (1 + tg_dec) / (wacc_dec - tg_dec)
    pv_terminal_value = terminal_value * forecast_df["discount_factor"].iloc[-1]

    enterprise_value = forecast_df["pv_fcff"].sum() + pv_terminal_value
    net_debt = df["debt"].iloc[-1] - df["cash_operating"].iloc[-1]
    equity_value = enterprise_value - net_debt

    # PRIMARY: shares outstanding read DIRECTLY (Titan has this column natively)
    shares_direct = df["equity_shares"].iloc[-1]

    # CROSS-CHECK: derive independently via Share Capital / Face Value,
    # same method Tata Motors required. Should closely match shares_direct.
    shares_derived = (df["equity_share_cap"].iloc[-1] * 1e7) / FACE_VALUE_RS
    diff_pct = abs(shares_direct - shares_derived) / shares_direct * 100
    if diff_pct > 1.0:
        log.warning(f"Direct shares ({shares_direct:,.0f}) vs derived "
                    f"({shares_derived:,.0f}) differ by {diff_pct:.2f}% - "
                    f"investigate before trusting per-share value.")
    else:
        log.info(f"Direct shares ({shares_direct:,.0f}) matches derived "
                  f"({shares_derived:,.0f}) within {diff_pct:.2f}% - cross-check passed.")

    value_per_share = (equity_value * 1e7) / shares_direct

    return {
        "enterprise_value_cr": enterprise_value, "terminal_value_cr": terminal_value,
        "pv_terminal_value_cr": pv_terminal_value, "net_debt_cr": net_debt,
        "equity_value_cr": equity_value, "shares_outstanding": shares_direct,
        "value_per_share": value_per_share,
    }


def main():
    checks = preflight_cross_check()

    df = pd.read_csv(PIPELINE_CSV)
    df = df.sort_values("year").reset_index(drop=True)

    wacc = checks["wacc_at_C9"]
    terminal_growth = checks["terminal_growth_C22"]

    all_results = []
    for scenario in ["bear", "base", "bull"]:
        forecast_df = project_financials(df, checks, scenario, terminal_growth)
        valuation = discount_and_value(forecast_df, wacc, terminal_growth, df, checks)

        log.info(f"[{scenario.upper()}] EV={valuation['enterprise_value_cr']:.0f}cr | "
                  f"Equity Value={valuation['equity_value_cr']:.0f}cr | "
                  f"Value/Share=Rs.{valuation['value_per_share']:.2f}")

        forecast_df.to_csv(f"data/titan_forecast_{scenario}.csv", index=False)
        all_results.append({"scenario": scenario, **valuation})

    out = pd.DataFrame(all_results)
    out.to_csv(OUTPUT_CSV, index=False)
    log.info(f"Saved valuation summary to {OUTPUT_CSV}")


if __name__ == "__main__":
    main()