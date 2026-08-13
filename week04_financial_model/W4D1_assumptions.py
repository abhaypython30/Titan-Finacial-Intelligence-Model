"""
Titan Company - Week 4, Day 1: Assumption Derivation & WACC

"""

import logging
import numpy as np
import pandas as pd
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

PIPELINE_CSV = "data/titan_pipeline_new.csv"
EXCEL_PATH = "data/Titan project.xlsx"
ASSUMPTIONS_SHEET = "Assumptions"
OUTPUT_CSV = "data/titan_assumptions_derived.csv"


COVID_YEARS = [2020, 2021]

# BETA SOURCED: Yahoo Finance, 5Y monthly, TITAN.NS, checked Aug 2026.
# (e.g. screener.in or moneycontrol) before treating as fully settled.
BETA = 0.18


def load_pipeline():
    df = pd.read_csv(PIPELINE_CSV)
    df = df.sort_values("year").reset_index(drop=True)
    log.info(f"Loaded {len(df)} rows, sorted by year: {df['year'].tolist()}")
    return df


def revenue_growth_stats(df):
    growth = df["sales"].pct_change().dropna() * 100
    stats = {
        "growth_10yr_avg": growth.mean(),
        "growth_last3yr_avg": growth.tail(3).mean(),
        "growth_median": growth.median(),
        "growth_std": growth.std(),
        "growth_bear_p25": growth.quantile(0.25),
        "growth_base_p50": growth.quantile(0.50),
        "growth_bull_p75": growth.quantile(0.75),
    }

    non_covid_growth = df[~df["year"].isin(COVID_YEARS)]["sales"].pct_change().dropna() * 100
    stats["growth_last3yr_avg_ex_covid"] = non_covid_growth.tail(3).mean()

    diff = abs(stats["growth_last3yr_avg"] - stats["growth_last3yr_avg_ex_covid"])
    log.info(f"Preliminary COVID check (last-3yr avg): incl.={stats['growth_last3yr_avg']:.2f}%, "
              f"excl.={stats['growth_last3yr_avg_ex_covid']:.2f}%, diff={diff:.2f}pp. "
              f"NOTE: for Titan's FY17-26 window, the last 3 years (2024-2026) don't "
              f"actually contain a COVID year - this comparison will likely show little "
              f"or no difference REGARDLESS of the exclusion decision. This is expected, "
              f"not a sign COVID didn't matter - the real, decision-relevant comparison "
              f"(full-distribution Bear/Base/Bull with vs without COVID) is done properly "
              f"in Day2, not here. Treat this as a rough glance only.")

    return stats


def ebitda_margin_stats(df):
    margin = (df["ebitda"] / df["sales"]) * 100
    stats = {
        "margin_10yr_avg": margin.mean(),
        "margin_last3yr_avg": margin.tail(3).mean(),
        "margin_median": margin.median(),
        "margin_std": margin.std(),
        "margin_bear_p25": margin.quantile(0.25),
        "margin_base_p50": margin.quantile(0.50),
        "margin_bull_p75": margin.quantile(0.75),
    }
    log.info(f"EBITDA margin range: {margin.min():.2f}% to {margin.max():.2f}% - "
              f"cross-check against Week3's SQL-derived benchmark before trusting.")
    return stats


def capex_and_nwc(df):
    """
    Same standard definitions as Tata Motors - not verifiable against
    Excel per Rule 10 since no dedicated CAPEX/NWC row exists in the
    raw source, same situation as Tata Motors had.
    """
    df["capex"] = df["fixed_assets"].diff() + df["depreciation"]
    df["nwc"] = df["debtors"] + df["inventory"] + df["cash_operating"] - df["payables"]
    df["dNWC"] = df["nwc"].diff()
    df["capex_pct_sales"] = df["capex"] / df["sales"]

    negative_capex_years = df.loc[df["capex"] < 0, "year"].tolist()
    if negative_capex_years:
        log.warning(f"Negative CAPEX found in: {negative_capex_years} - investigate "
                    f"before this feeds Week5.")
    else:
        log.info("CAPEX positive in all years - passed.")

    capex_pct_sales_fwd = df["capex_pct_sales"].tail(3).mean()
    return df, capex_pct_sales_fwd


def dso_dio_dpo(df):
    return {
        "dso_fwd": df["debtor_days"].tail(3).mean(),
        "dio_fwd": df["inventory_days"].tail(3).mean(),
        "dpo_fwd": df["payables_days"].tail(3).mean(),
    }


def tax_rate(df):
    profitable = df[df["ebt"] > 0]
    implied_rate = (profitable["tax"] / profitable["ebt"]).mean() * 100
    log.info(f"Implied effective tax rate: {implied_rate:.2f}%")
    return implied_rate


def kd_from_data(df):
    implied_kd = (df["interest"] / df["debt"]).mean() * 100
    log.info(f"Implied Kd (interest/debt): {implied_kd:.2f}%")
    return implied_kd


def compute_wacc(df, rf, beta, erp, kd, tax_pct):
    ke = rf + beta * erp
    kd_after_tax = kd * (1 - tax_pct / 100)

    equity_book = (df["reserves"] + df["equity_share_cap"]).iloc[-1]
    debt_book = df["debt"].iloc[-1]
    total = equity_book + debt_book

    e_weight = equity_book / total
    d_weight = debt_book / total
    wacc = e_weight * ke + d_weight * kd_after_tax

    log.info(f"Ke={ke:.2f}% | Kd(after-tax)={kd_after_tax:.2f}% | "
              f"E-weight={e_weight:.2%} | D-weight={d_weight:.2%} | WACC={wacc:.2f}%")
    return wacc, ke, kd_after_tax, e_weight, d_weight


def read_excel_assumptions(path, sheet):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    inputs = {
        "rf": ws["C3"].value,
        "beta": ws["C4"].value,
        "erp": ws["C5"].value,
        "kd": ws["C6"].value,
        "tax_for_kd": ws["C7"].value,
    }
    log.info(f"Read Assumptions tab inputs: {inputs}")
    return inputs, wb, ws


def write_results_to_excel(wb, ws, wacc, beta_used):
    ws["C4"] = beta_used  # confirm sourced Beta is written, not left blank
    ws["C9"] = round(wacc, 2)
    wb.save(EXCEL_PATH)
    log.info(f"Wrote WACC to {EXCEL_PATH} (C9), Beta confirmed at C4.")


def main():
    df = load_pipeline()

    excel_inputs, wb, ws = read_excel_assumptions(EXCEL_PATH, ASSUMPTIONS_SHEET)

    rev_stats = revenue_growth_stats(df)
    margin_stats = ebitda_margin_stats(df)
    df, capex_pct_sales_fwd = capex_and_nwc(df)
    wc_fwd = dso_dio_dpo(df)
    tax_pct = tax_rate(df)
    implied_kd = kd_from_data(df)

    kd_to_use = excel_inputs["kd"] if excel_inputs["kd"] else implied_kd
    tax_to_use = excel_inputs["tax_for_kd"] if excel_inputs["tax_for_kd"] else round(tax_pct, 2)

    wacc, ke, kd_after_tax, e_weight, d_weight = compute_wacc(
        df,
        rf=excel_inputs["rf"],
        beta=BETA,
        erp=excel_inputs["erp"],
        kd=kd_to_use,
        tax_pct=tax_to_use,
    )

    results = {
        **rev_stats, **margin_stats,
        "capex_pct_sales_fwd": capex_pct_sales_fwd, **wc_fwd,
        "tax_rate_implied": tax_pct, "kd_implied": implied_kd,
        "kd_used": kd_to_use, "tax_used": tax_to_use,  # NEW: explicit record of
                                                         # what actually drove WACC,
                                                         # vs the always-computed
                                                         # implied values above
        "rf": excel_inputs["rf"], "beta": BETA, "erp_used": excel_inputs["erp"],
        "ke": ke, "kd_after_tax": kd_after_tax,
        "e_weight": e_weight, "d_weight": d_weight, "wacc": wacc,
    }

    out_df = pd.DataFrame([results])
    out_df.to_csv(OUTPUT_CSV, index=False)
    log.info(f"Saved derived assumptions to {OUTPUT_CSV}")

    write_results_to_excel(wb, ws, wacc, BETA)
    df.to_csv(PIPELINE_CSV, index=False)
    log.info("Updated titan_pipeline_new.csv with capex, nwc, dNWC columns.")


if __name__ == "__main__":
    main()